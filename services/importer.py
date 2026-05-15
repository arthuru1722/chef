import csv
import io
import re
from datetime import date, datetime, time

from data.defaults import DEFAULT_VALUES
from data.fields import FIELD_GROUPS
from services.dates import normalize_text
from services.forms import empty_values


ALLOWED_SPREADSHEET_EXTENSIONS = {"csv", "xlsx"}
MATERIAL_COLUMNS = 10


def allowed_spreadsheet(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_SPREADSHEET_EXTENSIONS


def importable_field_keys():
    fields = []
    for _, group_fields in FIELD_GROUPS:
        for key, _ in group_fields:
            if DEFAULT_VALUES.get(key) in ("", [], None):
                fields.append(key)

    for index in range(1, MATERIAL_COLUMNS + 1):
        fields.extend([f"material{index}Nome", f"material{index}Custo"])

    return fields


def field_aliases():
    aliases = {}
    repeated_labels = set()

    for group, group_fields in FIELD_GROUPS:
        for key, label in group_fields:
            aliases[normalize_text(key)] = key
            aliases[normalize_text(f"{group} {label}")] = key
            normalized_label = normalize_text(label)
            if normalized_label in aliases:
                repeated_labels.add(normalized_label)
            else:
                aliases[normalized_label] = key

    for repeated in repeated_labels:
        aliases.pop(repeated, None)

    aliases.update({
        normalize_text("nome"): "nomeContratante",
        normalize_text("contratante"): "nomeContratante",
        normalize_text("dataEvento"): "__data_evento__",
        normalize_text("data da festa"): "__data_evento__",
        normalize_text("horaEvento"): "__hora_evento__",
        normalize_text("hora da festa"): "__hora_evento__",
        normalize_text("dataParcela1"): "__data_parcela_1__",
        normalize_text("data da 1a parcela"): "__data_parcela_1__",
        normalize_text("dataParcela2"): "__data_parcela_2__",
        normalize_text("data da 2a parcela"): "__data_parcela_2__",
        normalize_text("dataAssinatura"): "__data_assinatura__",
        normalize_text("data da assinatura"): "__data_assinatura__",
    })

    for index in range(1, MATERIAL_COLUMNS + 1):
        aliases[normalize_text(f"material{index}Nome")] = f"material{index}Nome"
        aliases[normalize_text(f"material{index}")] = f"material{index}Nome"
        aliases[normalize_text(f"material {index}")] = f"material{index}Nome"
        aliases[normalize_text(f"material{index}Custo")] = f"material{index}Custo"
        aliases[normalize_text(f"custo material {index}")] = f"material{index}Custo"

    return aliases


def cell_to_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_date(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)

    text = cell_to_text(value)
    for pattern in ("%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    return None


def parse_time(value):
    if isinstance(value, datetime):
        return value.hour, value.minute
    if isinstance(value, time):
        return value.hour, value.minute

    text = cell_to_text(value)
    match = re.search(r"(\d{1,2})[:hH](\d{2})", text)
    if match:
        return int(match.group(1)), int(match.group(2))
    if text.isdigit():
        return int(text), 0
    return None


def apply_split_date(values, prefix, value):
    parsed = parse_date(value)
    if not parsed:
        return
    values[f"{prefix}Dia"] = f"{parsed.day:02d}"
    values[f"{prefix}Mes"] = f"{parsed.month:02d}"
    values[f"{prefix}Ano"] = str(parsed.year)


def apply_signature_date(values, value):
    parsed = parse_date(value)
    if not parsed:
        return
    months = [
        "janeiro", "fevereiro", "março", "abril", "maio", "junho",
        "julho", "agosto", "setembro", "outubro", "novembro", "dezembro",
    ]
    values["dataFinalDia"] = f"{parsed.day:02d}"
    values["dataFinalMesExtenso"] = months[parsed.month - 1]
    values["dataFinalAnoUltimoDigito"] = str(parsed.year)[-1]


def apply_event_date(values, value):
    parsed = parse_date(value)
    if not parsed:
        return
    values["diaEvento"] = f"{parsed.day:02d}"
    values["mesEvento"] = f"{parsed.month:02d}"
    values["anoEvento"] = str(parsed.year)
    if parsed.hour or parsed.minute:
        values["horaEvento"] = f"{parsed.hour:02d}"
        values["minutoEvento"] = f"{parsed.minute:02d}"


def apply_event_time(values, value):
    parsed = parse_time(value)
    if not parsed:
        return
    hour, minute = parsed
    values["horaEvento"] = f"{hour:02d}"
    values["minutoEvento"] = f"{minute:02d}"


def apply_material(materials, target, value):
    match = re.match(r"material(\d+)(Nome|Custo)", target)
    if not match:
        return
    index = int(match.group(1)) - 1
    field = "nome" if match.group(2) == "Nome" else "custo"
    while len(materials) <= index:
        materials.append({"nome": "", "custo": ""})
    materials[index][field] = cell_to_text(value)


def apply_cell(values, materials, target, value):
    if not target:
        return
    if target == "__data_evento__":
        apply_event_date(values, value)
        return
    if target == "__hora_evento__":
        apply_event_time(values, value)
        return
    if target == "__data_parcela_1__":
        apply_split_date(values, "parcela1", value)
        return
    if target == "__data_parcela_2__":
        apply_split_date(values, "parcela2", value)
        return
    if target == "__data_assinatura__":
        apply_signature_date(values, value)
        return
    if target.startswith("material"):
        apply_material(materials, target, value)
        return
    values[target] = cell_to_text(value)


def parse_csv(file_storage):
    text = file_storage.stream.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return reader.fieldnames or [], list(reader)


def parse_xlsx(file_storage):
    from openpyxl import load_workbook

    workbook = load_workbook(file_storage.stream, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [cell_to_text(cell) for cell in rows[0]]
    data_rows = []
    for row in rows[1:]:
        data_rows.append({headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))})
    return headers, data_rows


def parse_spreadsheet(file_storage):
    extension = file_storage.filename.rsplit(".", 1)[1].lower()
    headers, rows = parse_csv(file_storage) if extension == "csv" else parse_xlsx(file_storage)

    aliases = field_aliases()
    mapped_headers = {header: aliases.get(normalize_text(header)) for header in headers}
    contracts = []

    for row in rows:
        values = empty_values()
        materials = []
        has_content = False
        for header, raw_value in row.items():
            if raw_value not in (None, ""):
                has_content = True
            apply_cell(values, materials, mapped_headers.get(header), raw_value)
        if has_content:
            if materials:
                values["materiais"] = [
                    material for material in materials if material.get("nome") or material.get("custo")
                ] or values["materiais"]
            contracts.append(values)

    return contracts


def build_template_workbook():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Contratos"

    headers = importable_field_keys()
    sheet.append(headers)
    sheet.append(example_row())

    fill = PatternFill("solid", fgColor="E7EEEC")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill

    for index, header in enumerate(headers, start=1):
        sheet.column_dimensions[cell_column(index)].width = max(14, min(32, len(header) + 2))

    return workbook


def build_template_csv():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(importable_field_keys())
    writer.writerow(example_row())
    return output.getvalue().encode("utf-8-sig")


def values_to_export_row(values):
    row = []
    materials = values.get("materiais") or []
    for header in importable_field_keys():
        material_match = re.match(r"material(\d+)(Nome|Custo)", header)
        if material_match:
            index = int(material_match.group(1)) - 1
            field = "nome" if material_match.group(2) == "Nome" else "custo"
            row.append(materials[index].get(field, "") if index < len(materials) else "")
        else:
            row.append(values.get(header, ""))
    return row


def build_contracts_csv(rows):
    from services.contracts import sorted_contract_rows, values_from_row

    row_by_id = {row["id"]: row for row in rows}
    output = io.StringIO()
    writer = csv.writer(output)
    headers = importable_field_keys()
    writer.writerow(headers)
    for item in sorted_contract_rows(rows):
        values = values_from_row(row_by_id[item["id"]])
        writer.writerow(values_to_export_row(values))
    return output.getvalue().encode("utf-8-sig")


def cell_column(index):
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def example_row():
    row = []
    examples = {
        "nomeContratante": "Maria Exemplo",
        "nacionalidadeContratante": "brasileira",
        "estadoCivilContratante": "solteira",
        "profissaoContratante": "autonoma",
        "rgContratante": "123456",
        "cpfContratante": "000.000.000-00",
        "diaEvento": "20",
        "mesEvento": "06",
        "anoEvento": "2026",
        "horaEvento": "18",
        "minutoEvento": "30",
        "tipoEvento": "Casamento",
        "qtdPessoas": "80",
        "parcela1Dia": "01",
        "parcela1Mes": "06",
        "parcela1Ano": "2026",
        "parcela2Dia": "10",
        "parcela2Mes": "06",
        "parcela2Ano": "2026",
        "dataFinalDia": "15",
        "dataFinalMesExtenso": "maio",
        "dataFinalAnoUltimoDigito": "6",
        "material1Nome": "Travessa",
        "material1Custo": "50,00",
    }
    for header in importable_field_keys():
        row.append(examples.get(header, ""))
    return row
